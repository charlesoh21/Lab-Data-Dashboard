import numpy
import pandas as pd
import openpyxl
import sqlite3
import os
import time
import configparser

#change workdir to where you keep your .db, .cfg, etc files
workdir = 'C:\\Users\\riley.w\\Documents\\SQLTest'


#used to get various env variables from .cfg file
def cfgLoad():
    os.chdir(workdir)
    cfg = configparser.ConfigParser()
    cfg.read('settings.cfg')
    return cfg

#THIS IS AN EXAMPLE SCRIPT OF CONVERTING AN .XLSX SHEET TO A .DB MYSQL DATABASE.
#A BIT RESOURCE INTENSIVE ON A LAPTOP CPU, BUT SHOULD BE ABLE TO OPTIMIZE AND RUN WELL
#ON A SERVER CPU.
def xlsx2db():

    #open our workbook and iterate through the worksheets we want to update. Make separate workbook for final data.
    writer = pd.ExcelWriter(str('formatted_excel_' + fileName), engine='xlsxwriter')
    for wsName in wsNames:
        if wsName == 'Aging Error Log':
            fileNameCur = fileNameLog
        else:
            fileNameCur = fileName
        wb = openpyxl.load_workbook(fileNameCur, data_only=True)
        print(wsName)
        ws = wb[wsName]
        
        #unmerge merged cells and replace the values of these cells with the merged cell value
        while ws.merged_cells:
            for cell_group in ws.merged_cells:
                val = str(cell_group.start_cell.value).strip()
                ws.unmerge_cells(str(cell_group))
                for merged_cell in cell_group.cells:
                    ws.cell(row=merged_cell[0], column=merged_cell[1]).value = val
        #print('cells unmerged!')

        #reset the formatting for EVERY cell
        for row in ws.iter_rows():
            for cell in row:
                cell.style = 'Normal'
        #print('cells normalized!')

        #for these cases, we need to reformat the headers since they are merged in the original excel sheet
        if wsName in ['Completed Aging D1y', 'Completed Aging D1z','Completed Aging D1x']:

            #iterate until we hit headers
            while ws.cell(1,1).value != 'Customer':
                ws.delete_rows(1, 1)

            #remove first row of unecessary headers
            ws.delete_rows(1,1)

            #merge specific headers for cols G-O
            for col in range(7, 16):
                ws.cell(2, col).value = str(ws.cell(1, col).value) + ' - ' + str(ws.cell(2, col).value)

            #remove second row of unecessary headers
            ws.delete_rows(1,1)

            #renaming headers to their correct names
            ws.cell(1, 4).value = str('Dimm Type')
            ws.cell(1, 5).value = str('Dimm Density')
        #print('cells headers formatted!')

        #get column names and row entries (data) from worksheet
        data = ws.values
        cols = list(next(data)[0:])
        data = list(data)

        #iterate through our column names to make sure we don't have any empty headers. If we do, we remove
        #the column of data entirely. Else, if there is a '/' in a header, we remove it to prevent later
        #formatting issues
        i = 0
        while i < len(cols):
            if cols[i] == None:
                for j in range(len(data)):
                    temp = list(data[j])
                    temp.pop(i)
                    data[j]  = temp
                cols.pop(i)
            elif '/' in cols[i]:
                cols[i] = cols[i].replace('/', '-')
                i += 1
            elif cols[i] in ['Part No.', 'Part Number']:
                for j in range(len(data)):
                    if list(data[j])[i] is not None and '-' in list(data[j])[i]:
                        head, sep, tail = list(data[j])[i].partition('-')
                        temp = list(data[j])
                        temp[i] = head
                        data[j] = temp
                i += 1
            else:
                i += 1

        #converts data/cols variables into dataframe
        df = pd.DataFrame(data, columns=cols)
        #print('df made!')

        #general special conditions to remove garbage data.
        df = df.dropna(how='all', axis=1) #get rid of empty cols.
        df = df.dropna(how='all', axis=0) #get rid of empty rows.
        if wsName == 'Aging Error Log':
            df = df[df['SN'].notnull()] #if serial number is empty, we assume row is garbage
            df = df.rename(columns={"Type": "Fail Type"}) #rename type to fail type to prevent issues in sharepoint
        if fileNameCur == fileNameLog:
            df = df[df['Server No'].notnull()] #if server column is empty, we assume this row is garbage, so we delete it
            df = df[df['Server No'] != ''] #if server column is empty, we assume this row is garbage, so we delete it
        else:
            df = df[df['Customer'] != 'Customer'] #if customer column is repeat header, we assume this row is garbage, so we delete it
            df = df[df['Customer'] != 'D1y'] #if customer column is repeat header, we assume this row is garbage, so we delete it
            df = df[df['Customer'].notnull()] #if customer column is empty, we assume this row is garbage, so we delete it
            df = df[df['Customer'] != ''] #if customer column is empty, we assume this row is garbage, so we delete it
        df = df.replace('-', '') #get rid of rows that are just a dash.
        df = df.replace('\n', ' - ', regex=True) #replace spaces ('\n') with a dash (' - ')
        con = sqlite3.connect(dbName) #make database file

        #try to convert from pandas dataframe to mysql db and add in_SP_list field to said db. 
        #If this fails then the table already exists, so we add each entry manually.
        try:
            df.to_sql(con=con, name=wsName, index=False, if_exists='fail')
            cur = con.cursor()
            cur.execute(str('ALTER TABLE \"' + wsName + '\" ADD in_SP_list BIT'))
        except:
            #connect to db and get all items
            cur = con.cursor()
            cur.execute("SELECT * FROM \"" + wsName + "\"")
            enumItems = cur.fetchall()  
            series = df.values
            #iterate through items from excel and see if they are already in DB. If they aren't, then we add them to the db
            if wsName == 'Aging Error Log':
                LogID = list(map(lambda x: x[4], enumItems))
                SN = list(map(lambda x: x[5], enumItems))
                for row in series:
                    if str(row[4]) in LogID and str(row[5]) in SN:
                        continue
                    else:
                        cur.execute('insert into \"' + wsName + '\" values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)', numpy.append(row, 0))
            else:
                Batch = list(map(lambda x: x[1], enumItems))
                for row in series:
                    if row[1] in Batch:
                        continue
                    else:
                        cur.execute('insert into \"' + wsName + '\" values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)', numpy.append(row, 0))
            con.commit()

        #write data to excel sheet
        cols = df.columns
        df.to_excel(writer, sheet_name=wsName, startrow=1, header=False, index=False)
        wsW = writer.sheets[wsName]
        for idx, val in enumerate(cols):
            wsW.write(0, idx, val)
    
    #save workbook and close
    writer.save()



if __name__ == '__main__':
    #setup workdir, get the time for the stopwatch
    os.chdir(workdir)
    T1 = time.time()

    #load settings
    settings = cfgLoad()
    dbName = settings.get("SETTINGS", 'dbName')
    wsNames = settings.get('SETTINGS', 'wsNames').split('\n')
    urlSP = settings.get('SETTINGS', 'urlSP')
    fileName = settings.get('SETTINGS', 'xlsxName')
    fileNameLog = settings.get('SETTINGS', 'xlsxNameLog')

    xlsx2db()

    #connect to db and print the column headers, as well as some example data
    con = sqlite3.connect(dbName)
    cur = con.cursor()
    for wsName in wsNames:
        cur.execute(str("SELECT * FROM \"" + wsName + "\" WHERE rowid in (1, 20)"))
        colNames = list(map(lambda x: x[0], cur.description))
        print(colNames)
        results = cur.fetchall()
        for line in results:
            print(line)

    #stop and print stopwatch
    T2 = time.time()
    print("\ntime elapsed: " + str(abs(T2 - T1)) + " seconds!\n")
