import numpy
import pandas as pd
import openpyxl
import sqlite3
import os
import configparser
import sqlite3
import SPlist

#change workdir to where you keep your .db, .cfg, etc files
workdir = 'C:\\Users\\user\\Documents\\PSTconfig'



#used to get various env variables from .cfg file
def cfgLoad() -> configparser.ConfigParser:
    '''reads the config file in the specified work directory and returns its contents.
    '''
    os.chdir(workdir)
    cfg = configparser.ConfigParser()
    cfg.read('settings.cfg')
    return cfg

def xlsx2db() -> dict:
    '''Converts all entries of multiple excel sheets / files defined in the cfg file to a .db sqlite database. 
    '''
    #load settings
    settings = cfgLoad()
    dbName = settings.get("SETTINGS", 'dbName')
    wsNames = settings.get('SETTINGS', 'wsNames').split('\n')
    fileName = settings.get('SETTINGS', 'xlsxName')
    fileNameLog = settings.get('SETTINGS', 'xlsxNameLog')

    print('start xlsx to db conversion...')

    #open our workbook and iterate through the worksheets we want to update. Make separate workbook for final data.
    dataframes = {}
    writer = pd.ExcelWriter(str('formatted_excel_' + fileName), engine='xlsxwriter')
    for wsName in wsNames:
        if wsName == 'Aging Error Log':
            fileNameCur = fileNameLog
        else:
            fileNameCur = fileName
        wb = openpyxl.load_workbook(fileNameCur, data_only=True)
        print(wsName)
        ws = wb[wsName]
        
        #unmerge merged cells and replace the values of these cells with the merged cell value
        while ws.merged_cells:
            for cell_group in ws.merged_cells:
                val = str(cell_group.start_cell.value).strip()
                ws.unmerge_cells(str(cell_group))
                for merged_cell in cell_group.cells:
                    ws.cell(row=merged_cell[0], column=merged_cell[1]).value = val
        #print('cells unmerged!')

        #reset the formatting for EVERY cell
        for row in ws.iter_rows():
            for cell in row:
                cell.style = 'Normal'
        #print('cells normalized!')

        #for these cases, we need to reformat the headers since they are merged in the original excel sheet
        if wsName in ['Completed Aging D1y', 'Completed Aging D1z','Completed Aging D1x']:

            #iterate until we hit headers
            while ws.cell(1,1).value != 'Customer':
                ws.delete_rows(1, 1)

            #remove first row of unecessary headers
            ws.delete_rows(1,1)

            #merge specific headers for cols G-O
            for col in range(7, 16):
                ws.cell(2, col).value = str(ws.cell(1, col).value) + ' - ' + str(ws.cell(2, col).value)

            #remove second row of unecessary headers
            ws.delete_rows(1,1)

            #renaming headers to their correct names
            ws.cell(1, 4).value = str('Dimm Type')
            ws.cell(1, 5).value = str('Dimm Density')
        #print('cells headers formatted!')

        #get column names and row entries (data) from worksheet
        data = ws.values
        cols = list(next(data)[0:])
        data = list(data)

        #iterate through our column names to make sure we don't have any empty headers. If we do, we remove
        #the column of data entirely. Else, if there is a '/' in a header, we remove it to prevent later
        #formatting issues
        i = 0
        while i < len(cols):
            if cols[i] == None:
                for j in range(len(data)):
                    temp = list(data[j])
                    temp.pop(i)
                    data[j]  = temp
                cols.pop(i)
            elif '/' in cols[i]:
                cols[i] = cols[i].replace('/', '-')
                i += 1
            elif cols[i] in ['Part No.', 'Part Number']:
                for j in range(len(data)):
                    if list(data[j])[i] is not None and '-' in list(data[j])[i]:
                        head, sep, tail = list(data[j])[i].partition('-')
                        temp = list(data[j])
                        temp[i] = head
                        data[j] = temp
                i += 1
            else:
                i += 1

        #converts data/cols variables into dataframe
        df = pd.DataFrame(data, columns=cols)
        #print('df made!')

        #general special conditions to remove garbage data.
        df = df.dropna(how='all', axis=1) #get rid of empty cols.
        df = df.dropna(how='all', axis=0) #get rid of empty rows.
        if wsName == 'Aging Error Log':
            df = df[df['SN'].notnull()] #if serial number is empty, we assume row is garbage
            df = df.rename(columns={"Type": "Fail Type"}) #rename type to fail type to prevent issues in sharepoint
        if fileNameCur == fileNameLog:
            df = df[df['Server No'].notnull()] #if server column is empty, we assume this row is garbage, so we delete it
            df = df[df['Server No'] != ''] #if server column is empty, we assume this row is garbage, so we delete it
        else:
            df = df[df['Customer'] != 'Customer'] #if customer column is repeat header, we assume this row is garbage, so we delete it
            df = df[df['Customer'] != 'D1y'] #if customer column is repeat header, we assume this row is garbage, so we delete it
            df = df[df['Customer'].notnull()] #if customer column is empty, we assume this row is garbage, so we delete it
            df = df[df['Customer'] != ''] #if customer column is empty, we assume this row is garbage, so we delete it
        df = df.replace('-', '') #get rid of rows that are just a dash.
        df = df.replace('\n', ' - ', regex=True) #replace spaces ('\n') with a dash (' - ')
        con = sqlite3.connect(dbName) #make database file

        #try to convert from pandas dataframe to mysql db and add in_SP_list field to said db. 
        #If this fails then the table already exists, so we add each entry manually.
        try:
            df.to_sql(con=con, name=wsName, index=False, if_exists='fail')
            cur = con.cursor()
            cur.execute(str('ALTER TABLE \"' + wsName + '\" ADD in_SP_list BIT'))
        except:
            #connect to db and get all items
            cur = con.cursor()
            cur.execute("SELECT * FROM \"" + wsName + "\"")
            enumItems = cur.fetchall()  
            series = df.values
            #iterate through items from excel and see if they are already in DB. If they aren't, then we add them to the db
            if wsName == 'Aging Error Log':
                LogID = list(map(lambda x: x[4], enumItems))
                SN = list(map(lambda x: x[5], enumItems))
                for row in series:
                    if str(row[4]) in LogID and str(row[5]) in SN:
                        continue
                    else:
                        cur.execute('insert into \"' + wsName + '\" values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)', numpy.append(row, 0))
            else:
                Batch = list(map(lambda x: x[1], enumItems))
                for row in series:
                    if row[1] in Batch:
                        continue
                    else:
                        cur.execute('insert into \"' + wsName + '\" values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)', numpy.append(row, 0))
            con.commit()

        #write data to excel sheet
        cols = df.columns
        df.to_excel(writer, sheet_name=wsName, startrow=1, header=False, index=False)
        wsW = writer.sheets[wsName]
        for idx, val in enumerate(cols):
            wsW.write(0, idx, val)

        dataframes[wsName] = df
    
    #save workbook and close
    writer.save()
    print('xlsx to db conversion done!')
    return dataframes

def db2dict(cur, row) -> dict:
    '''Row factory for sqlite3 connection. 
    Converts rows in database into a single dict.'''
    dict = {}
    for idx, col in enumerate(cur.description):
        dict[col[0]] = row[idx]
    return dict

def db2SPlist(dataframes: dict = None) -> None:
    '''converts local database entries to online Sharepoint list entries.'''

    print('start db to sp list conversion...')

    #setup and get connection to db
    con = sqlite3.connect(str(workdir + "\\AgingLogDB.db"))
    con.row_factory = db2dict
    cur = con.cursor()

    #get table names and put them in a list
    cur.execute("SELECT name FROM sqlite_master WHERE type='table';")
    enumTables = cur.fetchall()
    dbTableNames = []
    for tableName in enumTables:
        dbTableNames.append(tableName['name'])

    #iterate through tables and upload them to a SP list for each table
    for listName in dbTableNames:

        #get all entries and column names from current table
        curList = SPlist.SPlist(listName)
        cur.execute("SELECT * FROM \"" + listName + "\"")
        colNames = list(map(lambda x: x[0], cur.description))
        enumItems = cur.fetchall()

        #try to make a new list. If we succeed, set newList to true, else false.

        #if this is a newlist, initialize the column names.
        print('listName: ' + str(listName))
        print('newList: ' + str(curList.newList))
        if curList.newList:
            #rename default 'title' column to the first column in our db table
            curList.update_field_name(fieldNameOld='Title', fieldNameNew=colNames[0])

            #iterate through db cols
            for i in colNames:

                #inSPlist is a db entry used to determine whether an entry is already in SP. Thus, we skip adding a column for it to the SP list.
                if i == colNames[0] or i == "in_SP_list":
                    continue
                else:
                    #create the column and make it visible in our default view in sharepoint
                    curList.create_field(i)
                    curList.make_field_visible(i)
            
            #re-establish connection to list so that we get updated info, and make the default column not required
            r = curList.get_list()
            curList.no_require_field(colNames[0])

        #iterate through items in db and change their column names to internal sharepoint names
        for d in enumItems:
            for val in colNames:
                if val == 'in_SP_list':
                    continue
                elif val == 'Customer' or val== 'Server No':
                    d["Title"] = d.pop(val)
                else:
                    d[curList.fields[val]['name']] = d.pop(val)

        #iterate through items in DB table
        dupeCount = 0
        newCount = 0
        for item in enumItems:

            #if this is a newlist, we add all the items to the SP list
            if curList.newList:

                #change empty db entry to empty sp entry
                for k in item:
                    if item[k] == None:
                        item[k] = ''
                
                #remove in_SP_list entry since it isn't in SP list
                item.pop('in_SP_list')

                if listName == 'Aging Error Log' and len(str(item['DQ'])) >= 250:
                    item['DQ'] = item['DQ'][0:250]

                #upload entry
                newCount += 1
                response = curList.add_item_to_SP_list([item])

            #if item is marked as not being in SP, we do another check to see if it is a duplicate entry
            elif item['in_SP_list'] == None or int(item['in_SP_list']) == 0:

                #change empty db entry to empty sp entry
                for k in item:
                    if item[k] == None:
                        item[k] = ''
                
                #query the SP list to see if this is a dupe item by checking the Serial Number (SN) and Log ID
                query = {}
                fields = []
                if listName == 'Aging Error Log':
                    query={'Where': ['And', ('Eq', 'SN', item['SN']), ('Eq', 'Log ID', item['Log_x0020_ID'])]}
                    fields = ['SN', 'Log ID']
                    if '' in [item['SN'], item['Log_x0020_ID']]:
                        continue
                elif listName in  ['Completed Aging D1y', 'Completed Aging D1z','Completed Aging D1x']:
                    if item['Batch'] == '':
                        continue
                    query = {'Where': [('Eq', 'Batch', item['Batch'])]}
                    fields = ['Batch']


                #if query is a match, then entry is already in SP, so we continue
                response = curList.query_SP_list_items(fields=fields,query=query)
                if response :
                    dupeCount += 1
                    #print('dupe: ' + str(response))
                    continue

                #otherwise, we add the entry to SP
                else:  
                    newCount += 1
                    print('new item in SP! uploading...')
                    print('query: ' + str(query))
                    item.pop('in_SP_list')
                    curList.add_item_to_SP_list([item])
            else:
                dupeCount += 1
                continue

        #update all the in_SP_list entries in the db to true
        print('# of dupes: ' + str(dupeCount))
        print('# of new entries: ' + str(newCount))
        cur.execute("UPDATE \"" + listName +"\" SET in_SP_list = 1 WHERE in_SP_list isNull or in_SP_list = 0")
        con.commit()

    print('db to sp list conversion done!')
    


if __name__ == '__main__':
    db2SPlist(xlsx2db())